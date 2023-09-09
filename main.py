import openpyxl as xl
toexcel = xl.load_workbook('practise.xlsx')
tosheet = toexcel['Dinesh']
fromexcel = xl.load_workbook('from excel.xlsx')
fromsheet = fromexcel['Dinesh']
for i in range(2, tosheet.max_row+1):
    tem = tosheet.cell(i, 2)
    for j in range(2, fromsheet.max_row + 1):
        if tem.value == fromsheet.cell(j, 2).value:
            tosheet.cell(i, 14).value = fromsheet.cell(j, 24).value
for row in range(2, from_excel.max_row + 1):
    from_excel = from_excel.cell(row=row, column=9).value
    from_excel= from_excel.cell(row=row, column=10).value
   from_excel = from_excel.cell(row=row, column=11).value

    if from_excel is None and from_excel is None and from_excel is None:
        continue

    total_3_months_cons = 3(from_excel)

    from_excel.cell(row=row, column=13).value = total_3_months_cons
    from_excel.cell(row=1, column=13).value = "Total Cons 3 Months"

toexcel.save('new.xlsx')
